"""Parse MaxQuant proteinGroups.txt — and many MaxQuant-like variants —
into structured data for the frontend.

Robustness goals:
- Tolerate non-standard identity columns (e.g. ``Combined_locus_tag`` from
  custom pipelines) by falling back when ``Protein IDs`` / ``Majority protein
  IDs`` / ``Fasta headers`` are absent.
- Tolerate missing peptide-count columns by setting a sensible default
  (``peptides=1`` for every detected protein) so downstream filters do not
  drop everything.
- Distinguish per-sample quant columns (``Intensity Sample1``) from
  per-group aggregate columns (``Intensity GroupName`` where GroupName
  is also a prefix of one or more sample names) and exclude the latter
  from the sample list.
- Keep a quant type even if it is missing for a few samples — pad with
  zeros instead of dropping the entire type.
"""

import csv
import math
import re
from collections import defaultdict

# Per-sample column prefixes in MaxQuant proteinGroups.txt
QUANT_PREFIXES = [
    "MS/MS count",
    "LFQ intensity",
    "Intensity",
    "iBAQ",
    "Razor + unique peptides",
    "Unique peptides",
    "Peptides",
]

# Prefixes that have [%] after sample name
PCT_PREFIXES = [
    "Sequence coverage",
]

# Identity columns to try, in priority order
ID_COLUMNS = [
    "Protein IDs", "Majority protein IDs", "Combined_locus_tag",
    "Locus tag", "Locus", "Accession", "Protein ID", "ID", "id",
]

GENE_NAME_COLUMNS = [
    "Gene names", "Gene name", "Gene", "Symbol", "Protein name", "Protein names",
]


def _float(val):
    """Parse a cell as float. NaN, +/-Inf, and parse failures all collapse to 0.0
    so downstream JSON serialisation (Starlette rejects non-finite floats) and
    plotting (D3 scales blow up on Inf) stay safe."""
    try:
        f = float(val) if val else 0.0
    except (ValueError, TypeError):
        return 0.0
    return f if math.isfinite(f) else 0.0


def _int(val):
    """Parse a cell as int. Inf via overflow path is also coerced to 0."""
    try:
        return int(float(val)) if val else 0
    except (ValueError, TypeError, OverflowError):
        return 0


def _extract_gene_name(fasta_header):
    """Extract a short display name from the fasta header."""
    if not fasta_header:
        return ""
    first = fasta_header.split(";")[0].strip()
    parts = first.split(None, 1)
    if len(parts) > 1:
        return parts[1]
    return first


def _detect_samples(headers):
    """Detect sample names from per-sample quant column headers."""
    samples = set()
    for header in headers:
        for prefix in QUANT_PREFIXES:
            if header.startswith(prefix + " "):
                sample = header[len(prefix) + 1:]
                if sample and not sample.startswith("[") and not sample.startswith("("):
                    samples.add(sample)
    return sorted(samples)


def _filter_aggregate_samples(samples):
    """Drop names that look like per-group aggregates rather than per-sample.

    Heuristic: ``S`` is an aggregate if some other detected name ``S2`` starts
    with ``S`` followed by a separator ([space, underscore, hyphen, digit]) —
    e.g. drop ``"Se"`` when ``"Se M1"``, ``"Se M2"`` exist; drop ``"TM7x"``
    when ``"TM7x A1"`` exists. Returns (kept, aggregates).
    """
    aggregates = set()
    for s in samples:
        for s2 in samples:
            if s2 == s:
                continue
            if s2.startswith(s) and len(s2) > len(s):
                next_ch = s2[len(s)]
                if next_ch in (" ", "_", "-", ".") or next_ch.isdigit():
                    aggregates.add(s)
                    break
    return [s for s in samples if s not in aggregates], sorted(aggregates)


def _detect_quant_columns(headers, samples):
    """Map quantification types to their per-sample columns.

    Keep a quant type if AT LEAST 2 samples have a column for it (was: all).
    Missing samples are padded with zeros at parse time.
    """
    quant_columns = {}
    for prefix in QUANT_PREFIXES:
        cols = {}
        for sample in samples:
            col_name = f"{prefix} {sample}"
            if col_name in headers:
                cols[sample] = col_name
        if len(cols) >= 2:
            quant_columns[prefix] = cols

    for prefix in PCT_PREFIXES:
        cols = {}
        for sample in samples:
            for pattern in (f"{prefix} {sample} [%]", f"{prefix} {sample}"):
                if pattern in headers:
                    cols[sample] = pattern
                    break
        if len(cols) >= 2:
            quant_columns[f"{prefix} [%]"] = cols

    return quant_columns


def _auto_groups(samples):
    """Auto-detect sample groups by stripping the trailing replicate index.

    Replicate index = an optional separator (space/underscore/hyphen) followed
    by trailing digits. Letters are NOT consumed, so naming conventions that
    encode the condition with a letter prefix (e.g. ``Se M1`` vs ``Se S1``)
    correctly resolve into separate groups:

        "Sample1", "Sample2", "Sample3"               -> "Sample"
        "A1", "A2", "B1", "B2"                        -> "A", "B"
        "Se M1"…"Se M4", "Se S1"…"Se S4"              -> "Se M", "Se S"
        "TM7x A1"…"A4", "TM7x B1"…"B4"                -> "TM7x A", "TM7x B"
        "Ctrl_1"…"Ctrl_3", "Treated_1"…"Treated_3"    -> "Ctrl", "Treated"

    Also strips a separator + single trailing letter (``Un_A``→``Un``,
    ``Un_B``→``Un``) when no numeric replicate suffix was found. This is
    the common biological-replicate naming convention for RNA-seq and
    splits ``Un_A/B/C/D`` and ``At_A/B/C/D`` into ``Un`` and ``At`` groups.

    Falls back to leading-letters grouping when stripping yields an empty stem.
    """
    groups = defaultdict(list)
    for sample in samples:
        stem = re.sub(r"[\s_\-]?\d+\s*$", "", sample).rstrip(" _-")
        if stem == sample:
            # No numeric replicate suffix. Try a single-letter suffix
            # (``Un_A``→``Un``) but only when preceded by a separator so we
            # don't chop the trailing letter off plain words like ``Sample``.
            stem = re.sub(r"[\s_\-][A-Za-z]\s*$", "", sample).rstrip(" _-")
        if not stem:
            m = re.match(r"^([A-Za-z]+)", sample)
            stem = m.group(1) if m else "Ungrouped"
        groups[stem].append(sample)
    return dict(groups)


def _pick_first_present(row, candidates, default=""):
    """Return the first candidate column's value that is non-empty."""
    for c in candidates:
        v = row.get(c, "")
        if v:
            return v
    return default


def _detect_encoding(filepath):
    """Sniff for a BOM at the start of the file. Common case: Excel "Save As
    Tab-delimited Text" emits UTF-16 LE with BOM, which fails to parse as
    UTF-8. Returns the encoding name to pass to ``open()``."""
    try:
        with open(filepath, "rb") as f:
            head = f.read(4)
    except OSError:
        return "utf-8"
    if head[:2] == b"\xff\xfe":
        return "utf-16"        # UTF-16 LE BOM
    if head[:2] == b"\xfe\xff":
        return "utf-16"        # UTF-16 BE BOM
    if head[:3] == b"\xef\xbb\xbf":
        return "utf-8-sig"     # UTF-8 BOM — strip it transparently
    return "utf-8"


def _detect_delimiter(sample_text):
    """Pick the most likely delimiter from a sample of the first ~8 KB.

    Tries ``csv.Sniffer`` first (handles edge cases like quoted fields
    containing commas). Falls back to a count-based heuristic if Sniffer
    raises — preferring tab (proteinGroups.txt is tab-delimited and is
    the historical default), then comma, then semicolon, then pipe.
    """
    candidates = "\t,;|"
    # Sniffer needs at least one full line — skip it on tiny samples
    if len(sample_text) >= 16:
        try:
            return csv.Sniffer().sniff(sample_text, delimiters=candidates).delimiter
        except csv.Error:
            pass
    # Count-based fallback. Look at the first line only — header row should
    # contain the strongest delimiter signal regardless of data shape.
    first_line = sample_text.split("\n", 1)[0]
    counts = [(d, first_line.count(d)) for d in candidates]
    counts = [c for c in counts if c[1] > 0]
    if not counts:
        return "\t"  # Single-column file (unlikely useful but won't crash)
    counts.sort(key=lambda c: (-c[1], candidates.index(c[0])))
    return counts[0][0]


def _read_xlsx(filepath):
    """Read the FIRST sheet of an .xlsx workbook into ``(headers, rows)``.

    ``rows`` is a list of dicts mapping each header to the cell value as a
    string (so downstream parsers see the same shape as csv.DictReader).
    Empty cells become empty strings. Numeric cells are stringified with
    enough precision that ``float()`` round-trips exactly; this matches the
    behaviour the user gets when they "Save As .csv" from Excel.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ValueError(
            "Excel (.xlsx) upload requires the 'openpyxl' package. "
            "Reinstall pro-ker (pip install --upgrade proker) to pick it up."
        ) from e
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not read .xlsx file: {e}") from e
    try:
        ws = wb[wb.sheetnames[0]] if wb.sheetnames else None
        if ws is None:
            raise ValueError("Excel workbook has no sheets")
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], [], []
        # Strip trailing all-None columns (Excel often pads to a wider rectangle)
        headers = [("" if h is None else str(h)).strip() for h in header_row]
        while headers and headers[-1] == "":
            headers.pop()
        n_cols = len(headers)
        rows = []
        raw_rows = []
        for row in rows_iter:
            # Skip fully-blank rows (Excel files often have a blank trailing row)
            if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
                continue
            cells = [(""  if c is None else str(c)) for c in row[:n_cols]]
            # Pad if a row is shorter than the header (rare but possible)
            while len(cells) < n_cols:
                cells.append("")
            raw_rows.append(cells)
            rows.append({h: cells[i] for i, h in enumerate(headers) if h})
        return headers, rows, raw_rows
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _read_delimited(filepath):
    """Read a text file with auto-detected delimiter (tab/comma/semi/pipe)
    into ``(headers, rows, raw_rows)`` — same shape as ``_read_xlsx``."""
    csv.field_size_limit(10_000_000)
    encoding = _detect_encoding(filepath)
    # Sniff a sample to guess the delimiter
    try:
        with open(filepath, "r", encoding=encoding, errors="replace") as f:
            sample = f.read(8192)
    except OSError as e:
        raise ValueError(f"Could not read file: {e}") from e
    delim = _detect_delimiter(sample)
    with open(filepath, "r", encoding=encoding, errors="replace") as f:
        reader = csv.DictReader(f, delimiter=delim)
        headers = list(reader.fieldnames or [])
        if not headers:
            return [], [], []
        rows = []
        raw_rows = []
        for row in reader:
            raw_rows.append([(row.get(h, "") or "") for h in headers])
            rows.append(row)
        return headers, rows, raw_rows


def _read_table(filepath):
    """Single entry point: dispatch by file extension.

    Returns ``(headers, rows, raw_rows)`` where ``rows`` is a list of dicts
    (mirrors csv.DictReader shape) and ``raw_rows`` is a list-of-lists in
    header order suitable for round-trip export. Both parsers use this so
    ``.xlsx`` / ``.xls`` / ``.csv`` / ``.tsv`` / ``.txt`` and other delimited
    text files all flow through the same downstream code.
    """
    ext = ""
    try:
        from pathlib import Path as _P
        ext = _P(filepath).suffix.lower()
    except Exception:
        pass
    if ext in (".xlsx", ".xlsm"):
        return _read_xlsx(filepath)
    # .xls (legacy binary Excel) needs xlrd, which we don't ship — give a
    # clear error pointing the user at the easy fix.
    if ext == ".xls":
        raise ValueError(
            "Legacy .xls binary Excel files aren't supported. "
            "Open the file in Excel and Save As .xlsx (or .csv / .tsv), "
            "then upload that instead."
        )
    return _read_delimited(filepath)


def parse_protein_groups(filepath):
    """Parse a MaxQuant proteinGroups.txt (or variant) and return structured
    data. Accepts ``.txt`` / ``.tsv`` / ``.csv`` / other delimited text via
    auto-detected delimiter, plus ``.xlsx`` workbooks (first sheet)."""
    headers, rows, raw_rows = _read_table(filepath)
    if not headers:
        raise ValueError("Empty file or unrecognized format")

    header_set = set(headers)

    # Detect samples, then strip per-group aggregate columns
    raw_samples = _detect_samples(headers)
    if not raw_samples:
        raise ValueError(
            "Could not detect sample columns. "
            "Ensure the file has per-sample columns like 'Intensity A1'."
        )
    samples, dropped_aggregates = _filter_aggregate_samples(raw_samples)
    if not samples:
        raise ValueError(
            "Detected only per-group aggregate columns, no per-sample columns. "
            "Make sure your file has per-sample quant columns (e.g. 'Intensity A1')."
        )

    quant_columns = _detect_quant_columns(header_set, samples)
    if not quant_columns:
        raise ValueError("No quantification data columns detected.")

    groups = _auto_groups(samples)

    # Discover the best identity & gene-name columns to use
    id_col = next((c for c in ID_COLUMNS if c in header_set), None)
    majority_col = "Majority protein IDs" if "Majority protein IDs" in header_set else id_col
    gene_col = next((c for c in GENE_NAME_COLUMNS if c in header_set), None)
    fasta_col = "Fasta headers" if "Fasta headers" in header_set else None

    # File-level metadata: which standard columns are present
    has_peptides_col = any(c in header_set for c in
                           ("Peptides", "Razor + unique peptides", "Unique peptides"))
    has_contam_col = "Potential contaminant" in header_set
    has_reverse_col = "Reverse" in header_set
    has_byside_col = "Only identified by site" in header_set

    # Parse all rows
    proteins = []
    quant_data = {qt: {s: [] for s in samples} for qt in quant_columns}
    raw_headers = list(headers)

    contaminant_count = 0
    reverse_count = 0
    only_by_site_count = 0

    for row in rows:
        is_contaminant = row.get("Potential contaminant", "") == "+"
        is_reverse = row.get("Reverse", "") == "+"
        is_only_by_site = row.get("Only identified by site", "") == "+"

        if is_contaminant:
            contaminant_count += 1
        if is_reverse:
            reverse_count += 1
        if is_only_by_site:
            only_by_site_count += 1

        # Identity — fall back to the first non-empty column we know about
        protein_id = row.get(id_col, "") if id_col else ""
        majority_id = row.get(majority_col, "") if majority_col else protein_id
        if not majority_id:
            majority_id = protein_id
        fasta_header = row.get(fasta_col, "") if fasta_col else ""

        # Gene name: from dedicated column, or from fasta header
        if gene_col:
            gene_name = row.get(gene_col, "") or _extract_gene_name(fasta_header)
        else:
            gene_name = _extract_gene_name(fasta_header)

        # Peptide counts — read what's there; if NO peptide-count column
        # exists at all in the file, default to 1 so downstream filters
        # (e.g. peptides<1) don't drop every protein.
        if has_peptides_col:
            peptides = _int(row.get("Peptides", ""))
            unique_peptides = _int(row.get("Unique peptides", ""))
            razor = _int(row.get("Razor + unique peptides", ""))
        else:
            peptides = 1
            unique_peptides = 0
            razor = 0

        protein = {
            "id": protein_id,
            "majority_id": majority_id,
            "gene_name": gene_name,
            "fasta_header": fasta_header,
            "mol_weight": _float(row.get("Mol. weight [kDa]", "")),
            "sequence_length": _int(row.get("Sequence length", "")),
            "peptides": peptides,
            "unique_peptides": unique_peptides,
            "razor_unique_peptides": razor,
            "sequence_coverage": _float(row.get("Sequence coverage [%]", "")),
            "score": _float(row.get("Score", "")),
            "only_identified_by_site": is_only_by_site,
            "reverse": is_reverse,
            "potential_contaminant": is_contaminant,
            "peptide_sequences": row.get("Peptide sequences", ""),
            # Tag every protein with its source file for multi-file workflows.
            # main.py overwrites this with the user-facing filename.
            "source_file": str(filepath),
        }
        proteins.append(protein)

        # Collect quantification values; missing cols (sample present in
        # `samples` but missing for this quant type) are padded with 0.
        for qt, sample_cols in quant_columns.items():
            for sample in samples:
                col_name = sample_cols.get(sample)
                val = _float(row.get(col_name, "0")) if col_name else 0.0
                quant_data[qt][sample].append(val)

    # If the file had NO id column at all, synthesize one from the row index
    # so the UI has something to display.
    if not id_col:
        for idx, p in enumerate(proteins):
            p["id"] = f"row_{idx + 1}"
            p["majority_id"] = p["majority_id"] or f"row_{idx + 1}"

    return {
        "filename": str(filepath),
        "proteins": proteins,
        "samples": samples,
        "quant_types": list(quant_columns.keys()),
        "quant_data": quant_data,
        "suggested_groups": groups,
        "total_proteins": len(proteins),
        "contaminants": contaminant_count,
        "reverse_hits": reverse_count,
        "only_by_site": only_by_site_count,
        # Provenance / format flags so the frontend can adapt UI/filters
        "format_info": {
            "id_column": id_col,
            "gene_column": gene_col,
            "fasta_column": fasta_col,
            "has_peptide_counts": has_peptides_col,
            "has_contaminant_column": has_contam_col,
            "has_reverse_column": has_reverse_col,
            "has_only_by_site_column": has_byside_col,
            "dropped_aggregate_columns": dropped_aggregates,
        },
        # Original-file preservation for downstream Excel/CSV export
        "raw_headers": raw_headers,
        "raw_rows": raw_rows,
    }


# Sample-column detection for transcriptomics tables.
#
# Naive "exclude-everything-on-this-list" approaches are fragile because
# real-world RNA-seq output (featureCounts, htseq-count, salmon/kallisto,
# edgeR, DESeq2, limma, plus various per-pipeline glue scripts) introduces
# annotation and stats columns the static list never anticipated — gene
# coordinates, biotypes, transcript IDs, raw / scaled / shrunken estimates,
# per-comparison p-values, etc. So instead of a single hard list, we use
# two complementary signals:
#
#   1. NAME signal  — case-insensitively normalise the header (lower, drop
#      separators) and check against a comprehensive set of known metadata
#      names AND a list of substring patterns. The substring patterns catch
#      labelled stats columns like ``Adj.P.Val.GroupA_vs_B`` (which exact
#      match would miss).
#
#   2. VALUE signal — peek at the first ~40 rows. A column whose non-empty
#      cells are >= 80% numeric is a sample-counts CANDIDATE; a column with
#      strings (gene symbols, descriptions, chromosome names) is not.
#
# A column is treated as a sample iff the name signal does NOT mark it as
# annotation/stats AND the value signal says it's numeric. Both signals
# need to agree, so even a column with an unfamiliar name like
# ``Replicate_FB1`` will be picked up correctly (numeric values, name not
# on the exclusion list), and a column called ``log2_FoldChange.Cond1``
# will be correctly excluded (substring matches ``log2foldchange``).
#
# Exact-match metadata names (after lowercasing + replacing space/dot/hyphen
# with underscore + collapsing repeats). Single-letter entries like ``f``,
# ``t``, ``b``, ``p`` only fire on exact match — they would never appear as
# a sample column on their own anyway.
_RNA_META_NAMES = frozenset({
    # Identifiers
    "locustag", "locus_tag", "geneid", "gene_id", "id", "name",
    "transcript_id", "transcriptid", "ensemblid", "ensembl_id",
    "ensembl_gene_id", "ensembl_transcript_id", "refseq", "refseq_id",
    "uniprot", "uniprot_id", "entrez", "entrez_id", "ncbi_id",
    "symbol", "gene", "gene_name", "gene_symbol", "genename",
    "transcript_name", "transcriptname", "protein_id", "proteinid",
    # Description & annotation
    "description", "product", "function", "annotation", "note",
    "comment", "comments", "title",
    # Genomic coordinates
    "chr", "chrom", "chromosome", "seqname", "seqid",
    "start", "end", "strand", "length", "transcript_length",
    "gene_length", "exon_length", "cds_length",
    # Feature classification
    "featuretype", "feature_type", "feature", "type", "class",
    "biotype", "gene_biotype", "transcript_biotype", "category",
    # Summary statistics (per-row aggregates across samples — NOT samples
    # themselves; if these slipped through they'd skew the stats pipeline)
    "basemean", "base_mean", "mean", "median", "min", "max",
    "var", "variance", "sd", "stddev", "std_dev", "std",
    "se", "lfcse", "lfc_se", "lfc_standard_error",
    "ave_expr", "aveexpr", "avg_expr", "avg_expression",
    "average", "total", "sum",
    # Differential-expression statistics
    "f", "t", "b", "z", "stat", "statistic", "test_statistic",
    "fc", "foldchange", "fold_change",
    "logfc", "log2fc", "log2foldchange", "log_fc", "log2_fold_change",
    "logcpm", "log_cpm", "cpm", "rpm", "rpkm", "fpkm", "tpm",
    # P-values
    "p", "pvalue", "p_value", "pval", "p_val",
    # Adjusted P-values / FDR
    "padj", "p_adj", "adj_p", "adj_p_val", "adj_p_value",
    "fdr", "qvalue", "q_value", "q_val",
    "benjamini_hochberg_adjusted_pvalue", "bh_pvalue", "bh_p_value",
    "bonferroni",
})

# Substring patterns matched against the normalised name. Used to catch
# labelled stats columns the exact-match list can't (e.g. edgeR's
# ``logFC.GroupA_vs_GroupB`` becomes ``logfc_groupa_vs_groupb`` and matches
# the substring ``logfc``). Kept short so a sample called ``Sample_FDR1``
# doesn't accidentally get excluded — only patterns that are very unlikely
# to appear inside a sample identifier are listed.
_RNA_META_SUBSTRINGS = (
    "log2foldchange", "log2fc", "logfc",
    "log_cpm", "logcpm",
    "ave_expr", "aveexpr",
    "biotype", "ensembl",
)

# Word-bounded patterns: stats names that need to appear as a complete
# token (preceded by ``_`` or start, followed by ``_`` or end) so we don't
# false-positive on sample names like ``Sample_FDR1`` while still catching
# labelled stats columns like ``adj.P.Val.GroupA_vs_GroupB`` which
# normalises to ``adj_p_val_groupa_vs_groupb``. Stored as compiled regex
# at module-load time for speed.
import re as _re
_RNA_META_TOKEN_PATTERN = _re.compile(
    r"(^|_)("
    r"padj|p_value|pvalue|p_val|pval|"
    r"adj_p_val|adj_p_value|adj_p|"
    r"fdr|qvalue|q_value|q_val|"
    r"basemean|base_mean|"
    r"lfcse|lfc_se"
    r")(_|$)"
)


def _normalize_col_name(name):
    """Lower-case, replace space/dot/hyphen with underscore, collapse
    repeated underscores. Mirrors how the various tools format names so
    minor punctuation differences (``P.Value`` vs ``P_Value`` vs
    ``p value``) all collapse to the same canonical form."""
    if not name:
        return ""
    out = []
    prev_us = False
    for ch in str(name).strip().lower():
        if ch in (" ", "\t", ".", "-", "/"):
            if not prev_us:
                out.append("_")
                prev_us = True
        else:
            out.append(ch)
            prev_us = False
    return "".join(out).strip("_")


def _is_meta_column_by_name(name):
    """Return True if this column header looks like RNA-seq annotation or
    pre-computed stats (so it should NOT be treated as a sample). Three
    layers of matching, in order of strictness:

      1. Exact match against the comprehensive ``_RNA_META_NAMES`` set
         (after lower/separator-normalisation). Catches the canonical
         names every tool uses.
      2. Word-bounded regex against ``_RNA_META_TOKEN_PATTERN``. Catches
         labelled stats columns like ``adj.P.Val.GroupA_vs_B`` while
         leaving sample names like ``Sample_FDR1`` alone (the stats
         token must be a complete underscore-bounded segment).
      3. Plain substring against ``_RNA_META_SUBSTRINGS``. Catches the
         remaining patterns that are unique enough to never appear inside
         a real sample name (``logfc``, ``biotype``, ``ensembl``, etc.).
    """
    n = _normalize_col_name(name)
    if not n:
        return True  # Empty header — definitely not a sample column
    if n in _RNA_META_NAMES:
        return True
    if _RNA_META_TOKEN_PATTERN.search(n):
        return True
    for pat in _RNA_META_SUBSTRINGS:
        if pat in n:
            return True
    return False


def _column_numeric_score(rows, header, n_check=40):
    """Return (numeric_count, total_count) over the first ``n_check`` rows.

    Empty / NA / NaN / "." cells don't count toward either total — they're
    "missing", not "non-numeric". Only non-empty non-parseable strings
    count as a non-numeric.
    """
    n_total = 0
    n_numeric = 0
    NA_TOKENS = {"", "NA", "N/A", "NAN", "NULL", "NONE", ".", "-", "#N/A"}
    for row in rows[:n_check]:
        v = row.get(header, "")
        if v is None:
            continue
        s = str(v).strip()
        if s.upper() in NA_TOKENS:
            continue
        n_total += 1
        try:
            float(s)
            n_numeric += 1
        except (ValueError, TypeError):
            pass
    return n_numeric, n_total


def _detect_rna_samples(headers, rows):
    """Identify per-sample columns in an RNA-seq table.

    Returns ``(samples, exclusion_log)`` where ``exclusion_log`` is a list
    of ``(header, reason)`` tuples for every header that didn't make it.
    Reasons are surfaced via ``format_info.detection_log`` so the user can
    see why a column they expected to be a sample was filtered out.

    A header is treated as a sample iff:
      1. It's non-empty.
      2. The name doesn't match the metadata/stats list or substring
         patterns (``logFC``, ``P.Value``, ``Adj.P.Val.GroupA_vs_B``,
         ``Chr``, ``gene_biotype``, etc.).
      3. The first ~40 rows show >= 80% numeric values among non-empty
         cells. A column with text values can't be a sample-counts column
         no matter what its name says.
    """
    samples = []
    exclusion_log = []
    for h in headers:
        if not h:
            exclusion_log.append(("(empty header)", "empty"))
            continue
        if _is_meta_column_by_name(h):
            exclusion_log.append((h, "annotation/stats column name"))
            continue
        n_numeric, n_total = _column_numeric_score(rows, h)
        # Pure-empty column — can't tell if it's intended as a sample. Skip
        # it conservatively rather than raising; the user gets a 0-value
        # column in the output if they really wanted it included, but that's
        # easier to debug than silent exclusion.
        if n_total == 0:
            exclusion_log.append((h, "no values in first 40 rows"))
            continue
        if (n_numeric / n_total) < 0.8:
            exclusion_log.append(
                (h, f"non-numeric ({n_numeric}/{n_total} rows parse as float)"))
            continue
        samples.append(h)
    return samples, exclusion_log


def parse_transcriptomics(filepath):
    """Parse an RNA-seq differential-expression TSV (or similar tab-delimited
    transcriptomics output) into the same RAW shape that the frontend uses
    for proteomics. The columns we care about are the per-sample count
    columns — everything else is annotation or pre-computed statistics that
    the volcano-plot pipeline will recompute itself from the counts.

    Recognised non-count columns (skipped): Locustag, Gene, Description,
    FeatureType, logFC, PValue, Benjamini_Hochberg_Adjusted_PValue,
    plus a handful of common edgeR/DESeq2/limma equivalents.

    Every other column is assumed to be a per-sample count value. The user
    chose this format by toggling "Transcriptomics" in the upload panel —
    we don't try to be clever about format sniffing here; if the toggle
    misclassifies the file, the user gets a clear error from the upload
    flow rather than silent garbage.

    Counts are taken AS-IS — caller has indicated they're already
    normalised (the toggle's accompanying note states this) so no
    additional normalisation is applied. The dataset emerges with a
    single quant type called "Counts" so the frontend's processing UI
    stays simple.

    Accepts the same file formats as the proteomics parser: ``.txt`` /
    ``.tsv`` / ``.csv`` / other delimited text via auto-detected
    delimiter, plus ``.xlsx`` workbooks (first sheet).
    """
    headers, rows, _ = _read_table(filepath)
    if not headers:
        raise ValueError("Empty file or unrecognized format")

    # Detect sample columns via combined name+value classifier (see
    # _detect_rna_samples). This handles the long tail of RNA-seq tool
    # outputs that the simple "exclude these names" approach used to miss.
    samples, exclusion_log = _detect_rna_samples(headers, rows)
    if not samples:
        # Build a useful error pointing the user at WHY no samples were
        # detected — listing what was excluded helps them fix the file.
        excluded_summary = "; ".join(
            f"{h} ({reason})" for h, reason in exclusion_log[:10]
        )
        raise ValueError(
            "Could not detect sample count columns. "
            "Expected per-sample columns with numeric values alongside "
            "annotation columns (Locustag, Gene, Description, etc.).\n\n"
            f"Excluded columns: {excluded_summary}"
            + ("..." if len(exclusion_log) > 10 else "")
        )

    # Pick the gene-id and gene-name columns we'll use.
    id_col = next((c for c in ("Locustag", "locus_tag", "LocusTag", "GeneID",
                               "gene_id", "ID", "id") if c in headers), None)
    gene_col = next((c for c in ("Gene", "gene", "Symbol", "gene_name")
                     if c in headers), None)
    desc_col = next((c for c in ("Description", "description", "Product",
                                 "product") if c in headers), None)
    feat_col = next((c for c in ("FeatureType", "feature_type", "Feature",
                                 "Type") if c in headers), None)

    groups = _auto_groups(samples)

    proteins = []
    quant_data = {"Counts": {s: [] for s in samples}}
    raw_headers = list(headers)
    raw_rows = []

    for idx, row in enumerate(rows):
        raw_rows.append([(row.get(h, "") or "") for h in raw_headers])

        gene_id = row.get(id_col, "") if id_col else ""
        if not gene_id:
            gene_id = f"gene_{idx + 1}"
        gene_name = row.get(gene_col, "") if gene_col else ""
        description = row.get(desc_col, "") if desc_col else ""
        feature = row.get(feat_col, "") if feat_col else ""

        # Build a "protein" record. The frontend operates on this shape
        # (id, gene_name, peptides, etc.) so we keep the same keys; the
        # transcriptomics-specific bits (description, feature type) sit
        # alongside without the frontend caring.
        proteins.append({
            "id": gene_id,
            "majority_id": gene_id,
            "gene_name": gene_name,
            "fasta_header": description,  # Reuse for hover-tooltip text
            "mol_weight": 0.0,
            "sequence_length": 0,
            # peptides=1 keeps the default min-peptides filter (>=1) from
            # silently dropping every gene. Transcriptomics has no peptide
            # concept; the field is only relevant to MS data.
            "peptides": 1,
            "unique_peptides": 0,
            "razor_unique_peptides": 0,
            "sequence_coverage": 0.0,
            "score": 0.0,
            "only_identified_by_site": False,
            "reverse": False,
            "potential_contaminant": False,
            "peptide_sequences": "",
            "feature_type": feature,
            "description": description,
            "source_file": str(filepath),
        })

        for sample in samples:
            quant_data["Counts"][sample].append(_float(row.get(sample, "0")))

    return {
        "filename": str(filepath),
        "proteins": proteins,
        "samples": samples,
        "quant_types": ["Counts"],
        "quant_data": quant_data,
        "suggested_groups": groups,
        "total_proteins": len(proteins),
        "contaminants": 0,
        "reverse_hits": 0,
        "only_by_site": 0,
        "format_info": {
            "data_type": "transcriptomics",
            "id_column": id_col,
            "gene_column": gene_col,
            "description_column": desc_col,
            "feature_column": feat_col,
            "has_peptide_counts": False,
            "has_contaminant_column": False,
            "has_reverse_column": False,
            "has_only_by_site_column": False,
            "dropped_aggregate_columns": [],
            # Surface what got excluded from the sample list and why so the
            # user can spot misclassifications. Format: list of {column,
            # reason}. Truncated to first 20 entries to keep the JSON payload
            # bounded for files with many annotation columns.
            "non_sample_columns": [
                {"column": h, "reason": r} for h, r in exclusion_log[:20]
            ],
            "non_sample_columns_truncated": len(exclusion_log) > 20,
            # Hint the frontend that no normalisation should be applied — the
            # count values came in already normalised (RPM/TPM/CPM/DESeq2's
            # rlog/vst etc.) and re-normalising would compound the transform.
            "pre_normalized": True,
        },
        "raw_headers": raw_headers,
        "raw_rows": raw_rows,
    }
